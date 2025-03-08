import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime, date, timedelta
import calendar
from app.database import requests as db_req  # Модуль запросов к БД
import io
import os
import zipfile
import tempfile
from fpdf import FPDF
from aiogram.types import Message
from aiogram.types.input_file import FSInputFile

async def generate_report_excel(filters: dict) -> (BytesIO, str):
    """
    Генерирует профессиональный Excel‑отчёт по заказам с учетом фильтров.

    Параметры filters:
      - date: "all", "year", "month", "week" или период в формате "YYYY-MM-DD - YYYY-MM-DD"
      - manager: "all" или id менеджера
      - status: "all" или конкретный статус заказа

    Отчёт включает:
      • Заголовок с информацией о фильтрах (Период, Менеджер, Статус)
      • Таблицу с колонками:
          Номер заказа, Дата заказа, Менеджер, Товары, Статус, Сумма, Доставка, Оплата
      • Диаграмму (BarChart) с агрегированными данными по статусам заказов

    Возвращает кортеж (file_stream, filename)
    """
    # Получаем данные отчёта из БД
    report = await db_req.get_report_data(filters)
    orders = report.get("orders", [])

    # Определяем текст периода согласно фильтру date
    today = date.today()
    date_filter = filters.get("date", "all")
    if date_filter == "week":
        start_dt = today - timedelta(days=today.weekday())
        end_dt = start_dt + timedelta(days=6)
        period_text = f"за неделю ({start_dt} - {end_dt})"
    elif date_filter == "month":
        start_dt = today.replace(day=1)
        last_day = calendar.monthrange(today.year, today.month)[1]
        end_dt = today.replace(day=last_day)
        period_text = f"за месяц ({start_dt} - {end_dt})"
    elif date_filter == "year":
        start_dt = today.replace(month=1, day=1)
        end_dt = today.replace(month=12, day=31)
        period_text = f"за год ({start_dt} - {end_dt})"
    elif date_filter == "all":
        if orders:
            dates = []
            for order in orders:
                dt_str = order.get("order_datetime")
                try:
                    dt = datetime.strptime(dt_str, "%Y-%m-%d %H:%M:%S").date()
                    dates.append(dt)
                except Exception:
                    continue
            period_text = f"с {min(dates)} по {max(dates)}" if dates else "нет данных"
        else:
            period_text = "нет данных"
    else:
        try:
            start_str, end_str = date_filter.split(" - ")
            period_text = f"с {start_str} по {end_str}"
        except Exception:
            period_text = date_filter

    # Определяем описание менеджера.
    # Здесь вызываем функцию get_manager_fullname из manager_query.py
    manager_filter = filters.get("manager", "all")
    if manager_filter == "all":
        manager_text = "все менеджеры"
    else:
        manager_text = await db_req.get_manager_fullname(manager_filter)

    # Определяем описание статуса
    status_filter = filters.get("status", "all")
    status_text = "все статусы" if status_filter == "all" else status_filter

    # Создаем книгу и активный лист
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Отчет по заказам"

    # Определяем стили
    title_font = Font(size=18, bold=True, color="FFFFFF")
    title_fill = PatternFill("solid", fgColor="4F81BD")
    info_font = Font(italic=True, color="000000")
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="DCE6F1")
    align_center = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # Заголовок отчёта (объединяем ячейки от A1 до H1)
    ws.merge_cells("A1:H1")
    title_cell = ws["A1"]
    title_cell.value = "Отчет по заказам"
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = align_center

    # Строка с информацией о фильтрах (строка 2)
    ws.merge_cells("A2:C2")
    ws["A2"] = f"Период: {period_text}"
    ws["A2"].font = info_font
    ws["A2"].alignment = align_center

    ws.merge_cells("D2:F2")
    ws["D2"] = f"Менеджер: {manager_text}"
    ws["D2"].font = info_font
    ws["D2"].alignment = align_center

    ws.merge_cells("G2:H2")
    ws["G2"] = f"Статус: {status_text}"
    ws["G2"].font = info_font
    ws["G2"].alignment = align_center

    # Заголовки таблицы (начинаем с 4-й строки)
    columns = ["Номер заказа", "Дата заказа", "Менеджер", "Товары", "Статус", "Сумма", "Доставка", "Оплата"]
    header_row = 4
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = align_center
        cell.border = thin_border

    # Заполнение строк данными заказов (начинаем с 5-й строки)
    current_row = header_row + 1
    for order in orders:
        order_id = order.get("id")
        order_dt = order.get("order_datetime")
        manager_name = order.get("manager_name", "Не назначен")
        products_str = order.get("products", "")
        total_amount = order.get("total_amount", 0)
        status = order.get("status")
        delivery = order.get("delivery_method", "")
        payment = order.get("payment_method", "")

        row_values = [
            order_id, order_dt, manager_name,
            products_str, status, total_amount,
            delivery, payment
        ]
        for col_idx, val in enumerate(row_values, start=1):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.alignment = align_center
            cell.border = thin_border
        current_row += 1

    # Автоматическая подгонка ширины колонок
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_len + 4

    # Подготовка данных для диаграммы: количество заказов по статусам
    status_data = {}
    for order in orders:
        key = order.get("status", "Неизвестно")
        status_data[key] = status_data.get(key, 0) + 1

    # Создаем отдельный лист для данных диаграммы
    chart_sheet = wb.create_sheet("Диаграмма")
    chart_sheet.append(["Статус", "Количество"])
    for st, cnt in status_data.items():
        chart_sheet.append([st, cnt])

    # Формируем диаграмму
    chart = BarChart()
    chart.title = "Распределение заказов по статусу"
    chart.x_axis.title = "Статус"
    chart.y_axis.title = "Количество"
    chart.height = 10
    chart.width = 20

    data_ref = Reference(chart_sheet, min_col=2, min_row=1, max_row=chart_sheet.max_row)
    cats_ref = Reference(chart_sheet, min_col=1, min_row=2, max_row=chart_sheet.max_row)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)

    # Добавляем диаграмму на основной лист ниже таблицы
    ws.add_chart(chart, f"A{current_row + 2}")

    # Сохраняем книгу в поток и возвращаем его с именем файла
    output_stream = BytesIO()
    wb.save(output_stream)
    output_stream.seek(0)
    return output_stream, "order_report.xlsx"


async def generate_photo_id_pdf(message: Message) -> str:
    """
    Функция принимает сообщение с ZIP архивом, распаковывает его,
    для каждого фото отправляет его в Telegram для получения file_id,
    генерирует PDF, где на каждой странице фото и его file_id.
    Перед возвратом PDF удаляет все отправленные фото из чата.
    Возвращает путь к созданному PDF файлу и список file_id.
    """
    zip_file_bytes = io.BytesIO()
    file_info = await message.bot.get_file(message.document.file_id)
    await message.bot.download_file(file_info.file_path, destination=zip_file_bytes)
    zip_file_bytes.seek(0)

    file_ids = []
    results = []
    sent_messages = []  # Сохраняем объекты отправленных сообщений с фото

    try:
        with zipfile.ZipFile(zip_file_bytes) as zf:
            for file in zf.infolist():
                if file.filename.lower().endswith(('.png', '.jpg', '.jpeg')):
                    image_bytes = zf.read(file)
                    ext = os.path.splitext(file.filename)[1]
                    if ext.lower() not in ('.jpg', '.jpeg', '.png'):
                        ext = ".jpg"

                    with tempfile.NamedTemporaryFile(delete=False, suffix=ext) as temp_file:
                        temp_file.write(image_bytes)
                        temp_file.flush()
                        temp_path = temp_file.name

                    sent_msg = await message.bot.send_photo(
                        chat_id=message.chat.id,
                        photo=FSInputFile(temp_path)
                    )
                    os.remove(temp_path)

                    file_id = sent_msg.photo[-1].file_id
                    file_ids.append(file_id)
                    results.append((file.filename, image_bytes, file_id))
                    sent_messages.append(sent_msg)
    except Exception as e:
        raise Exception("Ошибка при обработке архива: " + str(e))

    pdf = FPDF()
    for filename, image_bytes, file_id in results:
        ext = os.path.splitext(filename)[1]
        if ext.lower() not in ('.jpg', '.jpeg', '.png'):
            ext = ".jpg"

        with tempfile.NamedTemporaryFile(delete=False, suffix=ext) as tmp_img:
            tmp_img.write(image_bytes)
            tmp_img.flush()
            tmp_img_path = tmp_img.name

        pdf.add_page()
        try:
            pdf.image(tmp_img_path, x=10, y=30, w=pdf.w - 20)
        except Exception as e:
            print(f"Ошибка при добавлении изображения {filename}: {e}")
        pdf.set_y(pdf.get_y() + 10)
        pdf.set_font("Arial", size=10)
        pdf.multi_cell(0, 10, file_id)
        os.remove(tmp_img_path)

    with tempfile.NamedTemporaryFile(delete=False, suffix=".pdf") as tmp_pdf:
        pdf_file_path = tmp_pdf.name
        pdf.output(pdf_file_path)

    # Удаляем все отправленные сообщения с фото из чата
    for sent_msg in sent_messages:
        try:
            await message.bot.delete_message(chat_id=message.chat.id, message_id=sent_msg.message_id)
        except Exception as e:
            print(f"Ошибка при удалении сообщения {sent_msg.message_id}: {e}")

    return pdf_file_path




