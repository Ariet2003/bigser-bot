import os
import re
from datetime import datetime

from aiogram.enums import ParseMode
from aiogram.types import Message, CallbackQuery, InlineKeyboardButton, InputMediaPhoto, InputMediaVideo
from aiogram.filters import CommandStart, Command
from aiogram import F, Router
from aiogram.types.input_file import BufferedInputFile, FSInputFile

from app.scripts import generate_report_excel, generate_photo_id_pdf
from app.users.admin import adminKeyboards as kb
from app.users.admin import adminStates as st
from app import utils
from app.database import requests as rq
from aiogram.fsm.context import FSMContext
from aiogram.types import InputFile
from app.utils import sent_message_add_screen_ids, router
from app import utils
from io import BytesIO
from openpyxl import Workbook, load_workbook


# Function to delete previous messages
async def delete_previous_messages(message: Message, telegram_id: str):
    # Проверяем, есть ли записи для этого пользователя
    if telegram_id not in sent_message_add_screen_ids:
        sent_message_add_screen_ids[telegram_id] = {'bot_messages': [], 'user_messages': []}

    user_data = sent_message_add_screen_ids[telegram_id]

    # Удаляем все сообщения пользователя, кроме "/start"
    for msg_id in user_data['user_messages']:
        try:
            if msg_id != message.message_id or message.text != "/start":
                await message.bot.delete_message(chat_id=telegram_id, message_id=msg_id)
        except Exception as e:
            print(f"Не удалось удалить сообщение {msg_id}: {e}")
    user_data['user_messages'].clear()

    # Удаляем все сообщения бота
    for msg_id in user_data['bot_messages']:
        try:
            if msg_id != message.message_id:
                await message.bot.delete_message(chat_id=telegram_id, message_id=msg_id)
        except Exception as e:
            print(f"Не удалось удалить сообщение {msg_id}: {e}")
    user_data['bot_messages'].clear()


# Administrator's personal account
async def admin_account(message: Message, state: FSMContext):
    tuid = message.chat.id
    if tuid not in sent_message_add_screen_ids:
        sent_message_add_screen_ids[tuid] = {'bot_messages': [], 'user_messages': []}
    user_data = sent_message_add_screen_ids[tuid]
    user_data['user_messages'].append(message.message_id)
    await delete_previous_messages(message, tuid)

    await state.clear()

    sent_message = await message.answer_photo(
        photo=utils.adminka_png,
        caption="Добро пожаловать в панель администратора! Выберите нужный раздел.",
        reply_markup=kb.admin_button
    )

    # Добавляем сообщение бота
    user_data['bot_messages'].append(sent_message.message_id)

# Хендлер для обработки команды "/photo"
# @router.message(Command("photo"))
# async def request_photo_handler(message: Message):
#     await message.answer("Пожалуйста, отправьте фото, чтобы я мог получить его ID.")
#
#
# # Хендлер для обработки фото от пользователя
# @router.message(F.photo)
# async def photo_handler(message: Message):
#     # Берем фотографию в самом большом разрешении и получаем ее ID
#     photo_id = message.photo[-1].file_id
#     await message.answer(f"ID вашей картинки: {photo_id}")


@router.callback_query(F.data == 'manage_employees')
async def manage_employees(callback_query: CallbackQuery, state: FSMContext):
    tuid = callback_query.message.chat.id
    user_data = sent_message_add_screen_ids[tuid]
    user_data['user_messages'].append(callback_query.message.message_id)

    await delete_previous_messages(callback_query.message, tuid)
    sent_message = await callback_query.message.answer_photo(photo=utils.adminka_png,
                                                             caption='Выберите нужное действие.',
                                                             reply_markup=kb.manage_employees_button)

    user_data['bot_messages'].append(sent_message.message_id)

@router.callback_query(F.data == 'go_to_dashboard')
async def go_to_dashboard(callback_query: CallbackQuery, state: FSMContext):
    await admin_account(callback_query.message, state)

@router.callback_query(F.data == 'add_employee')
async def add_employee(callback_query: CallbackQuery, state: FSMContext):
    tuid = callback_query.message.chat.id
    user_data = sent_message_add_screen_ids[tuid]
    user_data['user_messages'].append(callback_query.message.message_id)
    await delete_previous_messages(callback_query.message, tuid)

    sent_message = await callback_query.message.answer_photo(photo=utils.adminka_png,
                                                             caption='Выберите, какого сотрудника вы хотите добавить.',
                                                             reply_markup=kb.add_employee)

    user_data['bot_messages'].append(sent_message.message_id)

@router.callback_query(F.data == 'add_admin')
async def add_admin(callback_query: CallbackQuery, state: FSMContext):
    tuid = callback_query.message.chat.id
    user_data = sent_message_add_screen_ids[tuid]
    user_data['user_messages'].append(callback_query.message.message_id)
    await delete_previous_messages(callback_query.message, tuid)

    sent_message = await callback_query.message.answer(
        text="Напишите telegram id пользователя",
        reply_markup=kb.go_to_dashboard
    )
    await state.set_state(st.AddAdmin.write_telegram_id)

    user_data['bot_messages'].append(sent_message.message_id)

@router.message(st.AddAdmin.write_telegram_id)
async def add_admin_second(message: Message, state: FSMContext):
    tuid = message.chat.id
    user_data = sent_message_add_screen_ids[tuid]
    user_data['user_messages'].append(message.message_id)
    await delete_previous_messages(message, tuid)

    new_admin_telegram_id = message.text
    await state.update_data(new_admin_telegram_id=new_admin_telegram_id)
    sent_message = await message.answer(
        text="Напишите ФИО пользователя",
        reply_markup=kb.go_to_dashboard
    )

    await state.set_state(st.AddAdmin.write_fullname)
    user_data['bot_messages'].append(sent_message.message_id)


@router.message(st.AddAdmin.write_fullname)
async def add_admin_third(message: Message, state: FSMContext):
    tuid = message.chat.id
    user_data = sent_message_add_screen_ids[tuid]
    user_data['user_messages'].append(message.message_id)
    await delete_previous_messages(message, tuid)

    state_data = await state.get_data()
    new_admin_telegram_id = state_data.get('new_admin_telegram_id')
    new_admin_fullname = message.text

    print(new_admin_telegram_id)
    print(new_admin_fullname)

    is_added = await rq.add_or_update_user(telegram_id=new_admin_telegram_id,
                                     full_name=new_admin_fullname,
                                     role="ADMIN")

    if is_added:
        sent_message = await message.answer(
            text="Новый администратор успешно добавлен в систему.",
            reply_markup=kb.go_to_dashboard
        )
    else:
        sent_message = await message.answer(
            text="Не удалось добавить администратора. Пожалуйста, попробуйте снова.",
            reply_markup=kb.go_to_dashboard
        )

    user_data['bot_messages'].append(sent_message.message_id)

@router.callback_query(F.data == 'add_manager')
async def add_manager(callback_query: CallbackQuery, state: FSMContext):
    tuid = callback_query.message.chat.id
    user_data = sent_message_add_screen_ids[tuid]
    user_data['user_messages'].append(callback_query.message.message_id)
    await delete_previous_messages(callback_query.message, tuid)

    sent_message = await callback_query.message.answer(
        text="Напишите telegram id пользователя",
        reply_markup=kb.go_to_dashboard
    )
    await state.set_state(st.AddManager.write_telegram_id)

    user_data['bot_messages'].append(sent_message.message_id)

@router.message(st.AddManager.write_telegram_id)
async def add_manager_second(message: Message, state: FSMContext):
    tuid = message.chat.id
    user_data = sent_message_add_screen_ids[tuid]
    user_data['user_messages'].append(message.message_id)
    await delete_previous_messages(message, tuid)

    new_manager_telegram_id = message.text
    await state.update_data(new_manager_telegram_id=new_manager_telegram_id)
    sent_message = await message.answer(
        text="Напишите ФИО пользователя",
        reply_markup=kb.go_to_dashboard
    )

    await state.set_state(st.AddManager.write_fullname)
    user_data['bot_messages'].append(sent_message.message_id)


@router.message(st.AddManager.write_fullname)
async def add_manager_third(message: Message, state: FSMContext):
    tuid = message.chat.id
    user_data = sent_message_add_screen_ids[tuid]
    user_data['user_messages'].append(message.message_id)
    await delete_previous_messages(message, tuid)

    state_data = await state.get_data()
    new_manager_telegram_id = state_data.get('new_manager_telegram_id')
    new_manager_fullname = message.text

    print(new_manager_telegram_id)
    print(new_manager_fullname)

    is_added = await rq.add_or_update_user(telegram_id=new_manager_telegram_id,
                                     full_name=new_manager_fullname,
                                     role="MANAGER")

    if is_added:
        sent_message = await message.answer(
            text="Новый менеджер успешно добавлен в систему.",
            reply_markup=kb.go_to_dashboard
        )
    else:
        sent_message = await message.answer(
            text="Не удалось добавить менеджера. Пожалуйста, попробуйте снова.",
            reply_markup=kb.go_to_dashboard
        )

    user_data['bot_messages'].append(sent_message.message_id)


@router.callback_query(F.data == 'edit_employee')
async def edit_employee(callback_query: CallbackQuery, state: FSMContext):
    tuid = callback_query.message.chat.id
    user_data = sent_message_add_screen_ids[tuid]
    user_data['user_messages'].append(callback_query.message.message_id)
    await delete_previous_messages(callback_query.message, tuid)

    sent_message = await callback_query.message.answer_photo(photo=utils.adminka_png,
                                                             caption='Выберите, какого сотрудника вы хотите редактировать.',
                                                             reply_markup=kb.edit_employee)

    user_data['bot_messages'].append(sent_message.message_id)

@router.callback_query(F.data == "edit_admin")
async def handle_edit_admin(callback_query: CallbackQuery, state: FSMContext):
    page = 1  # начинаем с первой страницы
    admins = await rq.get_admins_by_page(page)
    total = await rq.get_total_admins()
    has_prev = page > 1
    has_next = (page * 10) < total
    markup = kb.create_admin_list_keyboard(admins, page, has_prev, has_next)

    # Если сообщение содержит фотографию, редактируем подпись, иначе – текст
    if callback_query.message.photo:
        await callback_query.message.edit_caption(caption="Список администраторов:", reply_markup=markup)
    else:
        await callback_query.message.edit_text(text="Список администраторов:", reply_markup=markup)


@router.callback_query(F.data.startswith("admin_page:"))
async def handle_admin_pagination(callback_query: CallbackQuery, state: FSMContext):
    page = int(callback_query.data.split(":")[1])
    admins = await rq.get_admins_by_page(page)
    total = await rq.get_total_admins()
    has_prev = page > 1
    has_next = (page * 10) < total
    markup = kb.create_admin_list_keyboard(admins, page, has_prev, has_next)
    await callback_query.message.edit_text(text="Список администраторов:", reply_markup=markup)


@router.callback_query(F.data.startswith("admin_detail:"))
async def handle_admin_detail(callback_query: CallbackQuery, state: FSMContext):
    admin_id = int(callback_query.data.split(":")[1])
    admin = await rq.get_admin_by_id(admin_id)
    if not admin:
        await callback_query.answer("Администратор не найден.")
        return
    text = f"Детали администратора:\nID: {admin['id']}\nФИО: {admin['full_name']}\nРоль: {admin['role']}"
    markup = kb.admin_detail_keyboard(admin_id)

    # Если сообщение содержит фотографию, изменяем подпись, иначе — текст.
    if callback_query.message.photo:
        await callback_query.message.edit_caption(caption=text, reply_markup=markup)
    else:
        await callback_query.message.edit_text(text=text, reply_markup=markup)


@router.callback_query(F.data.startswith("edit_admin_fullname:"))
async def edit_admin_fullname(callback_query: CallbackQuery, state: FSMContext):
    tuid = callback_query.message.chat.id
    user_data = sent_message_add_screen_ids[tuid]
    user_data['user_messages'].append(callback_query.message.message_id)
    await delete_previous_messages(callback_query.message, tuid)

    admin_id = int(callback_query.data.split(":")[1])
    await state.update_data(admin_id=admin_id)
    sent_message = await callback_query.message.answer("Введите новое ФИО для администратора:")
    await state.set_state(st.AdminEdit.waiting_for_fullname)

    user_data['bot_messages'].append(sent_message.message_id)

@router.message(st.AdminEdit.waiting_for_fullname)
async def process_admin_fullname(message: Message, state: FSMContext):
    tuid = message.chat.id
    user_data = sent_message_add_screen_ids[tuid]
    user_data['user_messages'].append(message.message_id)
    await delete_previous_messages(message, tuid)

    data = await state.get_data()
    admin_id = data.get("admin_id")
    new_fullname = message.text
    success = await rq.update_admin_fullname(admin_id, new_fullname)
    if success:
        sent_message = await message.answer("ФИО успешно изменено.", reply_markup=kb.go_to_dashboard)
        user_data['bot_messages'].append(sent_message.message_id)
    else:
        sent_message = await message.answer("Ошибка при изменении ФИО.", reply_markup=kb.go_to_dashboard)
        user_data['bot_messages'].append(sent_message.message_id)

    await state.clear()


@router.callback_query(F.data.startswith("edit_admin_role:"))
async def edit_admin_role(callback_query: CallbackQuery, state: FSMContext):
    tuid = callback_query.message.chat.id
    user_data = sent_message_add_screen_ids[tuid]
    user_data['user_messages'].append(callback_query.message.message_id)
    await delete_previous_messages(callback_query.message, tuid)

    admin_id = int(callback_query.data.split(":")[1])
    await state.update_data(admin_id=admin_id)
    sent_message = await callback_query.message.answer("Введите новую роль для администратора:")
    await state.set_state(st.AdminEdit.waiting_for_role)
    user_data['bot_messages'].append(sent_message.message_id)

@router.message(st.AdminEdit.waiting_for_role)
async def process_admin_role(message: Message, state: FSMContext):
    tuid = message.chat.id
    user_data = sent_message_add_screen_ids[tuid]
    user_data['user_messages'].append(message.message_id)
    await delete_previous_messages(message, tuid)

    data = await state.get_data()
    admin_id = data.get("admin_id")
    new_role = message.text
    success = await rq.update_admin_role(admin_id, new_role)
    if success:
        sent_message = await message.answer("Роль успешно изменена.", reply_markup=kb.go_to_dashboard)
        user_data['bot_messages'].append(sent_message.message_id)
    else:
        sent_message = await message.answer("Ошибка при изменении роли.", reply_markup=kb.go_to_dashboard)
        user_data['bot_messages'].append(sent_message.message_id)
    await state.clear()



@router.callback_query(F.data == "edit_manager")
async def handle_edit_manager(callback_query: CallbackQuery, state: FSMContext):
    page = 1  # начинаем с первой страницы
    managers = await rq.get_managers_by_page(page)
    total = await rq.get_total_managers()
    has_prev = page > 1
    has_next = (page * 10) < total
    markup = kb.create_manager_list_keyboard(managers, page, has_prev, has_next)

    # Если сообщение содержит фотографию, редактируем подпись, иначе – текст
    if callback_query.message.photo:
        await callback_query.message.edit_caption(caption="Список менеджеров:", reply_markup=markup)
    else:
        await callback_query.message.edit_text(text="Список менеджеров:", reply_markup=markup)


@router.callback_query(F.data.startswith("manager_page:"))
async def handle_manager_pagination(callback_query: CallbackQuery, state: FSMContext):
    page = int(callback_query.data.split(":")[1])
    managers = await rq.get_managers_by_page(page)
    total = await rq.get_total_managers()
    has_prev = page > 1
    has_next = (page * 10) < total
    markup = kb.create_manager_list_keyboard(managers, page, has_prev, has_next)
    await callback_query.message.edit_text(text="Список менеджеров:", reply_markup=markup)


@router.callback_query(F.data.startswith("manager_detail:"))
async def handle_manager_detail(callback_query: CallbackQuery, state: FSMContext):
    manager_id = int(callback_query.data.split(":")[1])
    manager = await rq.get_manager_by_id(manager_id)
    if not manager:
        await callback_query.answer("Менеджер не найден.")
        return
    text = f"Детали менеджера:\nID: {manager['id']}\nФИО: {manager['full_name']}\nРоль: {manager['role']}"
    markup = kb.manager_detail_keyboard(manager_id)

    # Если сообщение содержит фотографию, изменяем подпись, иначе — текст.
    if callback_query.message.photo:
        await callback_query.message.edit_caption(caption=text, reply_markup=markup)
    else:
        await callback_query.message.edit_text(text=text, reply_markup=markup)


@router.callback_query(F.data.startswith("edit_manager_fullname:"))
async def edit_manager_fullname(callback_query: CallbackQuery, state: FSMContext):
    tuid = callback_query.message.chat.id
    user_data = sent_message_add_screen_ids[tuid]
    user_data['user_messages'].append(callback_query.message.message_id)
    await delete_previous_messages(callback_query.message, tuid)

    manager_id = int(callback_query.data.split(":")[1])
    await state.update_data(manager_id=manager_id)
    sent_message = await callback_query.message.answer("Введите новое ФИО для менеджера:")
    await state.set_state(st.ManagerEdit.waiting_for_fullname)

    user_data['bot_messages'].append(sent_message.message_id)

@router.message(st.ManagerEdit.waiting_for_fullname)
async def process_manager_fullname(message: Message, state: FSMContext):
    tuid = message.chat.id
    user_data = sent_message_add_screen_ids[tuid]
    user_data['user_messages'].append(message.message_id)
    await delete_previous_messages(message, tuid)

    data = await state.get_data()
    manager_id = data.get("manager_id")
    new_fullname = message.text
    success = await rq.update_manager_fullname(manager_id, new_fullname)
    if success:
        sent_message = await message.answer("ФИО успешно изменено.", reply_markup=kb.go_to_dashboard)
        user_data['bot_messages'].append(sent_message.message_id)
    else:
        sent_message = await message.answer("Ошибка при изменении ФИО.", reply_markup=kb.go_to_dashboard)
        user_data['bot_messages'].append(sent_message.message_id)

    await state.clear()


@router.callback_query(F.data.startswith("edit_manager_role:"))
async def edit_manager_role(callback_query: CallbackQuery, state: FSMContext):
    tuid = callback_query.message.chat.id
    user_data = sent_message_add_screen_ids[tuid]
    user_data['user_messages'].append(callback_query.message.message_id)
    await delete_previous_messages(callback_query.message, tuid)

    manager_id = int(callback_query.data.split(":")[1])
    await state.update_data(manager_id=manager_id)
    sent_message = await callback_query.message.answer("Введите новую роль для менеджера:")
    await state.set_state(st.ManagerEdit.waiting_for_role)
    user_data['bot_messages'].append(sent_message.message_id)

@router.message(st.ManagerEdit.waiting_for_role)
async def process_manager_role(message: Message, state: FSMContext):
    tuid = message.chat.id
    user_data = sent_message_add_screen_ids[tuid]
    user_data['user_messages'].append(message.message_id)
    await delete_previous_messages(message, tuid)

    data = await state.get_data()
    manager_id = data.get("manager_id")
    new_role = message.text
    success = await rq.update_manager_role(manager_id, new_role)
    if success:
        sent_message = await message.answer("Роль успешно изменена.", reply_markup=kb.go_to_dashboard)
        user_data['bot_messages'].append(sent_message.message_id)
    else:
        sent_message = await message.answer("Ошибка при изменении роли.", reply_markup=kb.go_to_dashboard)
        user_data['bot_messages'].append(sent_message.message_id)
    await state.clear()


@router.callback_query(F.data == "delete_employee")
async def delete_employee(callback_query: CallbackQuery, state: FSMContext):
    text = "Выберите, какого сотрудника хотите удалить:"
    if callback_query.message.photo:
        await callback_query.message.edit_caption(
            caption=text,
            reply_markup=kb.delete_employee_role_keyboard
        )
    else:
        await callback_query.message.edit_text(
            text=text,
            reply_markup=kb.delete_employee_role_keyboard
        )


@router.callback_query(F.data == "delete_admin")
async def delete_admin(callback_query: CallbackQuery, state: FSMContext):
    page = 1
    admins = await rq.get_admins_by_page(page)
    total = await rq.get_total_admins()
    has_prev = page > 1
    has_next = (page * 10) < total
    markup = kb.create_delete_list_keyboard(admins, page, has_prev, has_next, role="admin")
    text = "Список администраторов:"
    if callback_query.message.photo:
        await callback_query.message.edit_caption(caption=text, reply_markup=markup)
    else:
        await callback_query.message.edit_text(text=text, reply_markup=markup)


@router.callback_query(F.data == "delete_manager")
async def delete_manager(callback_query: CallbackQuery, state: FSMContext):
    page = 1
    managers = await rq.get_managers_by_page(page)
    total = await rq.get_total_managers()
    has_prev = page > 1
    has_next = (page * 10) < total
    markup = kb.create_delete_list_keyboard(managers, page, has_prev, has_next, role="manager")
    text = "Список менеджеров:"
    if callback_query.message.photo:
        await callback_query.message.edit_caption(caption=text, reply_markup=markup)
    else:
        await callback_query.message.edit_text(text=text, reply_markup=markup)


@router.callback_query(F.data.startswith("delete_page:"))
async def handle_delete_pagination(callback_query: CallbackQuery, state: FSMContext):
    parts = callback_query.data.split(":")
    role = parts[1]
    page = int(parts[2])
    if role == "admin":
        employees = await rq.get_admins_by_page(page)
        total = await rq.get_total_admins()
        text = "Список администраторов:"
    elif role == "manager":
        employees = await rq.get_managers_by_page(page)
        total = await rq.get_total_managers()
        text = "Список менеджеров:"
    else:
        await callback_query.answer("Неверная роль.")
        return

    has_prev = page > 1
    has_next = (page * 10) < total
    markup = kb.create_delete_list_keyboard(employees, page, has_prev, has_next, role=role)
    if callback_query.message.photo:
        await callback_query.message.edit_caption(caption=text, reply_markup=markup)
    else:
        await callback_query.message.edit_text(text=text, reply_markup=markup)


@router.callback_query(F.data.startswith("delete_detail:"))
async def handle_delete_detail(callback_query: CallbackQuery, state: FSMContext):
    parts = callback_query.data.split(":")
    role = parts[1]
    user_id = int(parts[2])

    if role == "admin":
        employee = await rq.get_admin_by_id(user_id)
    elif role == "manager":
        employee = await rq.get_user_by_id(user_id)
    else:
        await callback_query.answer("Неверная роль.")
        return

    if not employee:
        await callback_query.answer("Сотрудник не найден.")
        return

    text = (
        f"Вы действительно хотите удалить сотрудника?\n\n"
        f"ID: {employee['id']}\nФИО: {employee['full_name']}\nРоль: {employee['role']}\n\n"
        "Внимание: При удалении сотрудника все его данные будут безвозвратно утрачены!"
    )
    markup = kb.confirm_delete_keyboard(employee['id'], role)
    if callback_query.message.photo:
        await callback_query.message.edit_caption(caption=text, reply_markup=markup)
    else:
        await callback_query.message.edit_text(text=text, reply_markup=markup)


@router.callback_query(F.data.startswith("confirm_delete:"))
async def confirm_delete(callback_query: CallbackQuery, state: FSMContext):
    tuid = callback_query.message.chat.id
    user_data = sent_message_add_screen_ids[tuid]
    user_data['user_messages'].append(callback_query.message.message_id)
    await delete_previous_messages(callback_query.message, tuid)

    parts = callback_query.data.split(":")
    role = parts[1]
    user_id = int(parts[2])
    confirm = parts[3]
    if confirm == "yes":
        success = await rq.delete_user_by_id(user_id)
        if success:
            sent_message = await callback_query.message.answer("Сотрудник успешно удалён.", reply_markup=kb.go_to_dashboard)
            user_data['bot_messages'].append(sent_message.message_id)
        else:
            sent_message = await callback_query.message.answer("Ошибка при удалении сотрудника.", reply_markup=kb.go_to_dashboard)
            user_data['bot_messages'].append(sent_message.message_id)
    else:
        sent_message = await callback_query.message.answer("Удаление отменено.", reply_markup=kb.go_to_dashboard)
        user_data['bot_messages'].append(sent_message.message_id)



@router.callback_query(F.data == 'manage_products')
async def manage_products(callback_query: CallbackQuery, state: FSMContext):
    tuid = callback_query.message.chat.id
    user_data = sent_message_add_screen_ids[tuid]
    user_data['user_messages'].append(callback_query.message.message_id)
    await delete_previous_messages(callback_query.message, tuid)

    sent_message = await callback_query.message.answer_photo(
        photo=utils.adminka_png,
        caption="Выберите нужное вам действие.",
        reply_markup=kb.manage_products_keyboard
    )

    user_data['bot_messages'].append(sent_message.message_id)


@router.callback_query(F.data == "manage_categories")
async def categories_menu(callback_query: CallbackQuery, state: FSMContext):
    tuid = callback_query.message.chat.id
    user_data = sent_message_add_screen_ids[tuid]
    user_data['user_messages'].append(callback_query.message.message_id)
    await delete_previous_messages(callback_query.message, tuid)

    sent_message = await callback_query.message.answer_photo(
        photo=utils.adminka_png,
        caption="Выберите нужное действие с категориями:",
        reply_markup=kb.categories_button
    )
    user_data['bot_messages'].append(sent_message.message_id)



@router.callback_query(F.data == "edit_categories")
async def edit_categories(callback_query: CallbackQuery, state: FSMContext):
    page = 1
    categories = await rq.get_categories_by_page(page)
    total = await rq.get_total_categories()
    has_prev = page > 1
    has_next = (page * 10) < total
    markup = kb.create_category_list_keyboard(categories, page, has_prev, has_next)
    text = "Список категорий:"
    if callback_query.message.photo:
        await callback_query.message.edit_caption(caption=text, reply_markup=markup)
    else:
        await callback_query.message.edit_text(text=text, reply_markup=markup)

@router.callback_query(F.data.startswith("category_page:"))
async def category_pagination(callback_query: CallbackQuery, state: FSMContext):
    page = int(callback_query.data.split(":")[1])
    categories = await rq.get_categories_by_page(page)
    total = await rq.get_total_categories()
    has_prev = page > 1
    has_next = (page * 10) < total
    markup = kb.create_category_list_keyboard(categories, page, has_prev, has_next)
    text = "Список категорий:"
    if callback_query.message.photo:
        await callback_query.message.edit_caption(caption=text, reply_markup=markup)
    else:
        await callback_query.message.edit_text(text=text, reply_markup=markup)


@router.callback_query(F.data.startswith("category_detail:"))
async def category_detail(callback_query: CallbackQuery, state: FSMContext):
    category_id = int(callback_query.data.split(":")[1])
    category = await rq.get_category_by_id(category_id)
    if not category:
        await callback_query.answer("Категория не найдена.")
        return
    text = f"Детали категории:\nID: {category['id']}\nНазвание: {category['name']}"
    markup = kb.category_detail_keyboard(category_id)
    if callback_query.message.photo:
        await callback_query.message.edit_caption(caption=text, reply_markup=markup)
    else:
        await callback_query.message.edit_text(text=text, reply_markup=markup)


@router.callback_query(F.data.startswith("edit_category:"))
async def edit_category(callback_query: CallbackQuery, state: FSMContext):
    tuid = callback_query.message.chat.id
    user_data = sent_message_add_screen_ids[tuid]
    user_data['user_messages'].append(callback_query.message.message_id)
    await delete_previous_messages(callback_query.message, tuid)

    category_id = int(callback_query.data.split(":")[1])
    await state.update_data(category_id=category_id)
    sent_message = await callback_query.message.answer("Введите новое название для категории:")
    await state.set_state(st.CategoryEdit.waiting_for_new_name)
    user_data['bot_messages'].append(sent_message.message_id)


@router.message(st.CategoryEdit.waiting_for_new_name)
async def process_category_edit(message: Message, state: FSMContext):
    tuid = message.chat.id
    user_data = sent_message_add_screen_ids[tuid]
    user_data['user_messages'].append(message.message_id)
    await delete_previous_messages(message, tuid)

    data = await state.get_data()
    category_id = data.get("category_id")
    new_name = message.text
    success = await rq.update_category_name(category_id, new_name)
    if success:
        sent_message = await message.answer("Название категории успешно изменено.", reply_markup=kb.go_to_dashboard)
        user_data['bot_messages'].append(sent_message.message_id)
    else:
        sent_message = await message.answer("Ошибка при изменении названия категории.", reply_markup=kb.go_to_dashboard)
        user_data['bot_messages'].append(sent_message.message_id)
    await state.clear()


@router.callback_query(F.data == "add_category")
async def add_category(callback_query: CallbackQuery, state: FSMContext):
    tuid = callback_query.message.chat.id
    user_data = sent_message_add_screen_ids[tuid]
    user_data['user_messages'].append(callback_query.message.message_id)
    await delete_previous_messages(callback_query.message, tuid)

    sent_message = await callback_query.message.answer("Введите название новой категории:",
                                                       reply_markup=kb.go_to_dashboard)

    user_data['bot_messages'].append(sent_message.message_id)
    await state.set_state(st.CategoryAdd.waiting_for_category_name)


@router.message(st.CategoryAdd.waiting_for_category_name)
async def process_category_add(message: Message, state: FSMContext):
    tuid = message.chat.id
    user_data = sent_message_add_screen_ids[tuid]
    user_data['user_messages'].append(message.message_id)
    await delete_previous_messages(message, tuid)

    category_name = message.text
    success = await rq.add_category(category_name)
    if success:
        sent_message = await message.answer("Новая категория успешно добавлена.", reply_markup=kb.go_to_dashboard)
        user_data['bot_messages'].append(sent_message.message_id)
    else:
        sent_message = await message.answer("Ошибка при добавлении категории.", reply_markup=kb.go_to_dashboard)
        user_data['bot_messages'].append(sent_message.message_id)
    await state.clear()


@router.callback_query(F.data == "delete_category")
async def delete_category_menu(callback_query: CallbackQuery, state: FSMContext):
    page = 1
    categories = await rq.get_categories_by_page(page)
    total = await rq.get_total_categories()
    has_prev = page > 1
    has_next = (page * 10) < total
    markup = kb.create_delete_category_list_keyboard(categories, page, has_prev, has_next)
    text = "Список категорий для удаления:"
    if callback_query.message.photo:
        await callback_query.message.edit_caption(caption=text, reply_markup=markup)
    else:
        await callback_query.message.edit_text(text=text, reply_markup=markup)


@router.callback_query(F.data.startswith("delete_category_page:"))
async def delete_category_pagination(callback_query: CallbackQuery, state: FSMContext):
    page = int(callback_query.data.split(":")[1])
    categories = await rq.get_categories_by_page(page)
    total = await rq.get_total_categories()
    has_prev = page > 1
    has_next = (page * 10) < total
    markup = kb.create_delete_category_list_keyboard(categories, page, has_prev, has_next)
    text = "Список категорий для удаления:"
    if callback_query.message.photo:
        await callback_query.message.edit_caption(caption=text, reply_markup=markup)
    else:
        await callback_query.message.edit_text(text=text, reply_markup=markup)


@router.callback_query(F.data.startswith("delete_category_detail:"))
async def delete_category_detail(callback_query: CallbackQuery, state: FSMContext):
    category_id = int(callback_query.data.split(":")[1])
    category = await rq.get_category_by_id(category_id)
    if not category:
        await callback_query.message.answer("Категория не найдена.")
        return
    text = (
        f"Вы действительно хотите удалить категорию?\n\n"
        f"ID: {category['id']}\nНазвание: {category['name']}\n\n"
        "Внимание: При удалении категория будет безвозвратно утрачена!"
    )
    markup = kb.confirm_delete_category_keyboard(category_id)
    if callback_query.message.photo:
        await callback_query.message.edit_caption(caption=text, reply_markup=markup)
    else:
        await callback_query.message.edit_text(text=text, reply_markup=markup)


@router.callback_query(F.data.startswith("confirm_delete_category:"))
async def confirm_delete_category(callback_query: CallbackQuery, state: FSMContext):
    tuid = callback_query.message.chat.id
    user_data = sent_message_add_screen_ids[tuid]
    user_data['user_messages'].append(callback_query.message.message_id)
    await delete_previous_messages(callback_query.message, tuid)

    parts = callback_query.data.split(":")
    category_id = int(parts[1])
    confirm = parts[2]
    if confirm == "yes":
        success = await rq.delete_category(category_id)
        if success:
            sent_message = await callback_query.message.answer("Категория успешно удалена.", reply_markup=kb.go_to_dashboard)
            user_data['bot_messages'].append(sent_message.message_id)
        else:
            sent_message = await callback_query.message.answer("Ошибка при удалении категории.", reply_markup=kb.go_to_dashboard)
            user_data['bot_messages'].append(sent_message.message_id)
    else:
        sent_message = await callback_query.message.answer("Удаление отменено.", reply_markup=kb.go_to_dashboard)
        user_data['bot_messages'].append(sent_message.message_id)


@router.callback_query(F.data == "manage_subcategories")
async def subcategories_menu(callback_query: CallbackQuery, state: FSMContext):
    """
    Отправляет сообщение с фото и клавиатурой с вариантами:
    - Редактировать подкатегории
    - Добавить подкатегорию
    - Удалить подкатегорию
    - ⬅️ Назад
    """
    tuid = callback_query.message.chat.id
    user_data = sent_message_add_screen_ids[tuid]
    user_data['user_messages'].append(callback_query.message.message_id)
    await delete_previous_messages(callback_query.message, tuid)

    sent_message = await callback_query.message.answer_photo(
        photo=utils.adminka_png,
        caption="Выберите нужное действие с подкатегориями:",
        reply_markup=kb.subcategories_button
    )
    user_data['bot_messages'].append(sent_message.message_id)


@router.callback_query(F.data == "edit_subcategories")
async def edit_subcategories(callback_query: CallbackQuery, state: FSMContext):
    """
    Выводит список подкатегорий для редактирования с пагинацией.
    """
    page = 1
    subcategories = await rq.get_subcategories_by_page(page)
    total = await rq.get_total_subcategories()
    has_prev = page > 1
    has_next = (page * 10) < total
    markup = kb.create_subcategory_list_keyboard(subcategories, page, has_prev, has_next)
    text = "Список подкатегорий:"
    if callback_query.message.photo:
        await callback_query.message.edit_caption(caption=text, reply_markup=markup)
    else:
        await callback_query.message.edit_text(text=text, reply_markup=markup)


@router.callback_query(F.data.startswith("subcategory_page:"))
async def subcategory_pagination(callback_query: CallbackQuery, state: FSMContext):
    """
    Обработчик пагинации списка подкатегорий.
    Callback_data: "subcategory_page:{page}"
    """
    page = int(callback_query.data.split(":")[1])
    subcategories = await rq.get_subcategories_by_page(page)
    total = await rq.get_total_subcategories()
    has_prev = page > 1
    has_next = (page * 10) < total
    markup = kb.create_subcategory_list_keyboard(subcategories, page, has_prev, has_next)
    text = "Список подкатегорий:"
    if callback_query.message.photo:
        await callback_query.message.edit_caption(caption=text, reply_markup=markup)
    else:
        await callback_query.message.edit_text(text=text, reply_markup=markup)


@router.callback_query(F.data.startswith("subcategory_detail:"))
async def subcategory_detail(callback_query: CallbackQuery, state: FSMContext):
    """
    При выборе подкатегории выводит её подробности.
    Callback_data: "subcategory_detail:{subcategory_id}"
    """
    subcategory_id = int(callback_query.data.split(":")[1])
    subcat = await rq.get_subcategory_by_id(subcategory_id)
    if not subcat:
        await callback_query.message.answer("Подкатегория не найдена.")
        return
    text = (
        f"Детали подкатегории:\n"
        f"ID: {subcat['id']}\nНазвание: {subcat['name']}\n"
        f"Категория: {subcat['category_name']}"
    )
    markup = kb.subcategory_detail_keyboard(subcategory_id)
    if callback_query.message.photo:
        await callback_query.message.edit_caption(caption=text, reply_markup=markup)
    else:
        await callback_query.message.edit_text(text=text, reply_markup=markup)


@router.callback_query(F.data.startswith("edit_subcategory:"))
async def edit_subcategory(callback_query: CallbackQuery, state: FSMContext):
    """
    Запускает процесс редактирования подкатегории.
    Сохраняет id подкатегории и запрашивает новое название (опционально).
    Callback_data: "edit_subcategory:{subcategory_id}"
    """
    tuid = callback_query.message.chat.id
    user_data = sent_message_add_screen_ids[tuid]
    user_data['user_messages'].append(callback_query.message.message_id)
    await delete_previous_messages(callback_query.message, tuid)

    subcategory_id = int(callback_query.data.split(":")[1])
    await state.update_data(subcategory_id=subcategory_id)
    sent_message = await callback_query.message.answer("Введите новое название подкатегории (оставьте пустым, если не меняете):")
    user_data['bot_messages'].append(sent_message.message_id)
    await state.set_state(st.SubcategoryEdit.waiting_for_new_name)


@router.message(st.SubcategoryEdit.waiting_for_new_name)
async def process_subcategory_new_name(message: Message, state: FSMContext):
    """
    Сохраняет новое название и запрашивает выбор родительской категории (с пагинацией).
    """
    tuid = message.chat.id
    user_data = sent_message_add_screen_ids[tuid]
    user_data['user_messages'].append(message.message_id)
    await delete_previous_messages(message, tuid)

    new_name = message.text.strip()
    await state.update_data(new_name=new_name)
    # Вывод списка категорий для выбора родительской
    page = 1
    categories = await rq.get_categories_by_page(page)
    total = await rq.get_total_categories()
    has_prev = page > 1
    has_next = (page * 10) < total
    markup = kb.create_parent_category_keyboard(categories, page, has_prev, has_next)
    sent_message = await message.answer("Выберите родительскую категорию для подкатегории:", reply_markup=markup)
    await state.set_state(st.SubcategoryEdit.waiting_for_parent_category)
    user_data['bot_messages'].append(sent_message.message_id)


@router.callback_query(F.data.startswith("parent_category:"))
async def parent_category_selected(callback_query: CallbackQuery, state: FSMContext):
    """
    Сохраняет выбранную родительскую категорию и выполняет обновление подкатегории.
    Callback_data: "parent_category:{category_id}"
    """
    tuid = callback_query.message.chat.id
    user_data = sent_message_add_screen_ids[tuid]
    user_data['user_messages'].append(callback_query.message.message_id)
    await delete_previous_messages(callback_query.message, tuid)

    parent_category_id = int(callback_query.data.split(":")[1])
    await state.update_data(parent_category_id=parent_category_id)
    data = await state.get_data()
    subcategory_id = data.get("subcategory_id")
    new_name = data.get("new_name")
    success = await rq.update_subcategory(subcategory_id, new_name, parent_category_id)
    if success:
        sent_message = await callback_query.message.answer("Подкатегория успешно изменена.", reply_markup=kb.go_to_dashboard)
        user_data['bot_messages'].append(sent_message.message_id)
    else:
        sent_message = await callback_query.message.answer("Ошибка при изменении подкатегории.", reply_markup=kb.go_to_dashboard)
        user_data['bot_messages'].append(sent_message.message_id)
    await state.clear()


@router.callback_query(F.data == "add_subcategory")
async def add_subcategory(callback_query: CallbackQuery, state: FSMContext):
    """
    Запускает процесс добавления новой подкатегории.
    """
    tuid = callback_query.message.chat.id
    user_data = sent_message_add_screen_ids[tuid]
    user_data['user_messages'].append(callback_query.message.message_id)
    await delete_previous_messages(callback_query.message, tuid)

    sent_message = await callback_query.message.answer("Введите название новой подкатегории:", reply_markup=kb.go_to_dashboard)
    user_data['bot_messages'].append(sent_message.message_id)
    await state.set_state(st.SubcategoryAdd.waiting_for_name)


@router.message(st.SubcategoryAdd.waiting_for_name)
async def process_subcategory_add_name(message: Message, state: FSMContext):
    """
    Сохраняет название новой подкатегории и запрашивает выбор родительской категории.
    """
    tuid = message.chat.id
    user_data = sent_message_add_screen_ids[tuid]
    user_data['user_messages'].append(message.message_id)
    await delete_previous_messages(message, tuid)

    new_name = message.text.strip()
    await state.update_data(new_name=new_name)
    page = 1
    categories = await rq.get_categories_by_page(page)
    total = await rq.get_total_categories()
    has_prev = page > 1
    has_next = (page * 10) < total
    markup = kb.create_parent_category_keyboard_add(categories, page, has_prev, has_next)
    sent_message = await message.answer("Выберите родительскую категорию для новой подкатегории:", reply_markup=markup)
    await state.set_state(st.SubcategoryAdd.waiting_for_parent_category)
    user_data['bot_messages'].append(sent_message.message_id)


@router.callback_query(F.data.startswith("parent_category_add:"))
async def parent_category_add(callback_query: CallbackQuery, state: FSMContext):
    """
    Сохраняет выбранную категорию и создает новую подкатегорию.
    Callback_data: "parent_category_add:{category_id}"
    """
    tuid = callback_query.message.chat.id
    user_data = sent_message_add_screen_ids[tuid]
    user_data['user_messages'].append(callback_query.message.message_id)
    await delete_previous_messages(callback_query.message, tuid)

    parent_category_id = int(callback_query.data.split(":")[1])
    await state.update_data(parent_category_id=parent_category_id)
    data = await state.get_data()
    new_name = data.get("new_name")
    success = await rq.add_subcategory(new_name, parent_category_id)
    if success:
        sent_message = await callback_query.message.answer("Новая подкатегория успешно добавлена.", reply_markup=kb.go_to_dashboard)
        user_data['bot_messages'].append(sent_message.message_id)
    else:
        sent_message = await callback_query.message.answer("Ошибка при добавлении подкатегории.", reply_markup=kb.go_to_dashboard)
        user_data['bot_messages'].append(sent_message.message_id)
    await state.clear()


@router.callback_query(F.data == "delete_subcategory")
async def delete_subcategory_menu(callback_query: CallbackQuery, state: FSMContext):
    """
    Выводит список подкатегорий для удаления с пагинацией.
    """
    page = 1
    subcategories = await rq.get_subcategories_by_page(page)
    total = await rq.get_total_subcategories()
    has_prev = page > 1
    has_next = (page * 10) < total
    markup = kb.create_delete_subcategory_list_keyboard(subcategories, page, has_prev, has_next)
    text = "Список подкатегорий для удаления:"
    if callback_query.message.photo:
        await callback_query.message.edit_caption(caption=text, reply_markup=markup)
    else:
        await callback_query.message.edit_text(text=text, reply_markup=markup)


@router.callback_query(F.data.startswith("delete_subcategory_page:"))
async def delete_subcategory_pagination(callback_query: CallbackQuery, state: FSMContext):
    """
    Обработчик пагинации для удаления подкатегорий.
    Callback_data: "delete_subcategory_page:{page}"
    """
    page = int(callback_query.data.split(":")[1])
    subcategories = await rq.get_subcategories_by_page(page)
    total = await rq.get_total_subcategories()
    has_prev = page > 1
    has_next = (page * 10) < total
    markup = kb.create_delete_subcategory_list_keyboard(subcategories, page, has_prev, has_next)
    text = "Список подкатегорий для удаления:"
    if callback_query.message.photo:
        await callback_query.message.edit_caption(caption=text, reply_markup=markup)
    else:
        await callback_query.message.edit_text(text=text, reply_markup=markup)


@router.callback_query(F.data.startswith("delete_subcategory_detail:"))
async def delete_subcategory_detail(callback_query: CallbackQuery, state: FSMContext):
    """
    При выборе подкатегории для удаления выводит подтверждение.
    Callback_data: "delete_subcategory_detail:{subcategory_id}"
    """

    subcategory_id = int(callback_query.data.split(":")[1])
    subcat = await rq.get_subcategory_by_id(subcategory_id)
    if not subcat:
        await callback_query.message.answer("Подкатегория не найдена.")
        return
    text = (
        f"Вы действительно хотите удалить подкатегорию?\n\n"
        f"ID: {subcat['id']}\nНазвание: {subcat['name']}\n"
        f"Родительская категория: {subcat['category_name']}\n\n"
        "Внимание: При удалении данные будут безвозвратно утрачены!"
    )
    markup = kb.confirm_delete_subcategory_keyboard(subcategory_id)
    if callback_query.message.photo:
        await callback_query.message.edit_caption(caption=text, reply_markup=markup)
    else:
        await callback_query.message.edit_text(text=text, reply_markup=markup)


@router.callback_query(F.data.startswith("confirm_delete_subcategory:"))
async def confirm_delete_subcategory(callback_query: CallbackQuery, state: FSMContext):
    """
    Обрабатывает подтверждение удаления подкатегории.
    Callback_data: "confirm_delete_subcategory:{subcategory_id}:{confirm}"
    """
    tuid = callback_query.message.chat.id
    user_data = sent_message_add_screen_ids[tuid]
    user_data['user_messages'].append(callback_query.message.message_id)
    await delete_previous_messages(callback_query.message, tuid)
    parts = callback_query.data.split(":")
    subcategory_id = int(parts[1])
    confirm = parts[2]
    if confirm == "yes":
        success = await rq.delete_subcategory(subcategory_id)
        if success:
            sent_message = await callback_query.message.answer("Подкатегория успешно удалена.", reply_markup=kb.go_to_dashboard)
            user_data['bot_messages'].append(sent_message.message_id)
        else:
            sent_message = await callback_query.message.answer("Ошибка при удалении подкатегории.", reply_markup=kb.go_to_dashboard)
            user_data['bot_messages'].append(sent_message.message_id)
    else:
        sent_message = await callback_query.message.answer("Удаление отменено.", reply_markup=kb.go_to_dashboard)
        user_data['bot_messages'].append(sent_message.message_id)


# Меню "Товары"
@router.callback_query(F.data == "manage_one_product")
async def products_menu(callback_query: CallbackQuery, state: FSMContext):
    tuid = callback_query.message.chat.id
    user_data = sent_message_add_screen_ids.setdefault(tuid, {"bot_messages": [], "user_messages": []})
    user_data["user_messages"].append(callback_query.message.message_id)
    await delete_previous_messages(callback_query.message, tuid)

    sent_message = await callback_query.message.answer_photo(
        photo=utils.adminka_png,
        caption="Если нужно внести изменения в товары, скачай файл, внесите необходимые добавления, изменения или "
                "удаления, затем отправь файл обратно для обновления товаров.",
        reply_markup=kb.product_menu_keyboard
    )
    user_data["bot_messages"].append(sent_message.message_id)


# Обработчик для экспорта данных о товарах в Excel
@router.callback_query(F.data == "edit_product")
async def edit_product_handler(callback_query: CallbackQuery, state: FSMContext):
    """
    При нажатии на кнопку "Редактировать товар" бот:
    1. Получает данные о товарах, цветах, размерах и подкатегориях из БД.
    2. Формирует Excel-файл с четырьмя вкладками:
       - "Товары": данные о товарах.
       - "Цвета": справочные данные по цветам.
       - "Размер": справочные данные по размерам.
       - "Подкатегория": список подкатегорий с информацией о родительской категории.
    3. Отправляет файл пользователю с инструкцией для редактирования.
    """
    tuid = callback_query.message.chat.id
    user_data = sent_message_add_screen_ids[tuid]
    user_data['user_messages'].append(callback_query.message.message_id)
    await delete_previous_messages(callback_query.message, tuid)

    # Получаем данные из БД
    products = await rq.get_all_products()
    colors = await rq.get_all_colors()
    sizes = await rq.get_all_sizes()
    subcategories = await rq.get_all_subcategories()

    # Создаем Excel-файл с вкладками
    wb = Workbook()
    # Вкладка "Товары"
    ws_products = wb.active
    ws_products.title = "Товары"
    headers = ["ID", "Название", "Цена", "Цветы", "Размеры", "Описание", "Тип продукта", "Материал", "Особенности",
               "Использование", "Температурный диапазон", "ID подкатегории"]
    ws_products.append(headers)
    for product in products:
        row = [
            product.get("id"),
            product.get("name"),
            product.get("price"),
            ", ".join(map(str, product.get("color_ids", []))) if product.get("color_ids") else "",
            ", ".join(map(str, product.get("size_ids", []))) if product.get("size_ids") else "",
            product.get("description"),
            product.get("product_type"),
            product.get("material"),
            product.get("features"),
            product.get("usage"),
            product.get("temperature_range"),
            product.get("subcategory_id")
        ]
        ws_products.append(row)

    # Вкладка "Цвета"
    ws_colors = wb.create_sheet("Цвета")
    ws_colors.append(["ID", "Название"])
    for color in colors:
        ws_colors.append([color.get("id"), color.get("name")])

    # Вкладка "Размер"
    ws_sizes = wb.create_sheet("Размер")
    ws_sizes.append(["ID", "Размер"])
    for size in sizes:
        ws_sizes.append([size.get("id"), size.get("size")])

    # Вкладка "Подкатегория"
    ws_subcat = wb.create_sheet("Подкатегория")
    ws_subcat.append(["ID", "Название", "ID родительской категории", "Название родительской категории"])
    for subcat in subcategories:
        ws_subcat.append([
            subcat.get("id"),
            subcat.get("name"),
            subcat.get("category_id"),
            subcat.get("category_name")
        ])

    # Сохраняем книгу в память
    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    # Создаем объект BufferedInputFile, передавая байты файла и имя
    input_file = BufferedInputFile(file_stream.getvalue(), filename="products.xlsx")

    caption_text = (
        "Скачайте таблицу для редактирования товаров.\n\n"
        "Вкладка 'Товары' содержит данные о товарах.\n"
        "Вкладки 'Цвета', 'Размер' и 'Подкатегория' предоставляют справочную информацию.\n\n"
        "После редактирования отправьте файл боту для обновления данных.\n\n"
        "Если хотите отменить обновление, отправьте '-' вместо файла."
    )

    sent_message = await callback_query.message.answer_document(
        document=input_file,
        caption=caption_text
    )
    user_data['bot_messages'].append(sent_message.message_id)

    # Переводим бота в состояние ожидания измененного Excel-файла
    await state.set_state(st.ProductEdit.waiting_for_excel_file.state)


# Новый обработчик для отмены обновления, если пользователь отправляет '-' вместо файла
@router.message(st.ProductEdit.waiting_for_excel_file, F.text)
async def cancel_excel_update(message: Message, state: FSMContext):
    if message.text.strip() == "-":
        await state.clear()
        await admin_account(message, state)
        return


# Обработчик для получения отредактированного Excel‑файла с товарами
@router.message(st.ProductEdit.waiting_for_excel_file, F.document)
async def process_edited_products(message: Message, state: FSMContext):
    """
    Обрабатывает полученный от пользователя Excel‑файл с измененными данными о товарах.
    Считывает лист "Товары" и для каждой строки:
      - Если указан ID и товар существует, обновляет данные (только если они изменились).
      - Если ID отсутствует или некорректен, добавляет новый товар.
    После обработки удаляет товары, которые отсутствуют в файле.
    По окончании отправляет результат операции.
    """
    tuid = message.chat.id
    user_data = sent_message_add_screen_ids[tuid]
    user_data['user_messages'].append(message.message_id)
    await delete_previous_messages(message, tuid)

    # Загружаем документ в память
    file_stream = BytesIO()
    bot = message.bot
    file_id = message.document.file_id
    file_info = await bot.get_file(file_id)
    await bot.download_file(file_info.file_path, destination=file_stream)
    file_stream.seek(0)

    try:
        wb = load_workbook(filename=file_stream, data_only=True)
    except Exception as e:
        sent_message = await message.answer(f"Ошибка при открытии файла: {e}",
                                            reply_markup=kb.go_to_dashboard)
        user_data['bot_messages'].append(sent_message.message_id)
        return

    if "Товары" not in wb.sheetnames:
        sent_message = await message.answer("В файле отсутствует вкладка 'Товары'. Проверьте файл и попробуйте снова.",
                                            reply_markup=kb.go_to_dashboard)
        user_data['bot_messages'].append(sent_message.message_id)
        return

    ws = wb["Товары"]
    rows = list(ws.rows)
    if not rows or len(rows) < 2:
        sent_message = await message.answer("Файл не содержит данных для обработки.",
                                            reply_markup=kb.go_to_dashboard)
        user_data['bot_messages'].append(sent_message.message_id)
        return

    # Получаем заголовки из первой строки
    headers = [cell.value for cell in rows[0]]
    header_map = {header: idx for idx, header in enumerate(headers) if header is not None}
    required_headers = ["ID", "Название", "Цена", "Цветы", "Размеры", "Описание", "Тип продукта",
                        "Материал", "Особенности", "Использование", "Температурный диапазон", "ID подкатегории"]
    for rh in required_headers:
        if rh not in header_map:
            sent_message = await message.answer(f"В файле отсутствует обязательный столбец: {rh}",
                                                reply_markup=kb.go_to_dashboard)
            user_data['bot_messages'].append(sent_message.message_id)
            return

    updated_count = 0
    added_count = 0
    deleted_count = 0

    # Вспомогательная функция для парсинга списков ID из строки
    def parse_ids(value):
        if value is None:
            return []
        if isinstance(value, str):
            return [int(x.strip()) for x in value.split(",") if x.strip().isdigit()]
        if isinstance(value, (int, float)):
            return [int(value)]
        return []

    # Вспомогательная функция для сравнения текущих данных товара с новыми
    def is_product_different(existing_product, new_data: dict) -> bool:
        fields = ["name", "price", "color_ids", "size_ids", "description", "product_type", "material",
                  "features", "usage", "temperature_range", "subcategory_id"]
        for field in fields:
            existing_value = getattr(existing_product, field, None)
            new_value = new_data.get(field)
            if field == "price":
                try:
                    if float(existing_value) != float(new_value):
                        return True
                except:
                    if existing_value != new_value:
                        return True
            elif field in ["color_ids", "size_ids"]:
                if list(existing_value or []) != new_value:
                    return True
            else:
                if existing_value != new_value:
                    return True
        return False

    # Собираем множество ID, представленных в Excel (только для строк с указанным валидным ID)
    excel_ids = set()

    # Обрабатываем каждую строку, начиная со второй (данные)
    for row in rows[1:]:
        # Читаем значения ячеек по заголовкам
        product_id = row[header_map["ID"]].value
        name = row[header_map["Название"]].value
        price = row[header_map["Цена"]].value
        color_ids_str = row[header_map["Цветы"]].value
        size_ids_str = row[header_map["Размеры"]].value
        description = row[header_map["Описание"]].value
        product_type = row[header_map["Тип продукта"]].value
        material = row[header_map["Материал"]].value
        features = row[header_map["Особенности"]].value
        usage = row[header_map["Использование"]].value
        temperature_range = row[header_map["Температурный диапазон"]].value
        subcategory_id = row[header_map["ID подкатегории"]].value

        try:
            price = float(price)
        except Exception:
            price = 0.0

        color_ids = parse_ids(color_ids_str)
        size_ids = parse_ids(size_ids_str)

        product_data = {
            "name": name,
            "price": price,
            "color_ids": color_ids,
            "size_ids": size_ids,
            "description": description,
            "product_type": product_type,
            "material": material,
            "features": features,
            "usage": usage,
            "temperature_range": temperature_range,
            "subcategory_id": int(subcategory_id) if subcategory_id is not None else None
        }

        if product_id is not None:
            try:
                product_id_int = int(product_id)
                excel_ids.add(product_id_int)
            except Exception:
                product_id_int = None

            if product_id_int:
                existing_product = await rq.get_product_by_id(product_id_int)
                if existing_product:
                    # Обновляем только если данные отличаются
                    if is_product_different(existing_product, product_data):
                        success = await rq.update_product(product_id_int, product_data)
                        if success:
                            updated_count += 1
                    # Если данные не изменились, обновление не производится
                else:
                    success = await rq.add_product(product_data)
                    if success:
                        added_count += 1
            else:
                success = await rq.add_product(product_data)
                if success:
                    added_count += 1
        else:
            # Если ID не указан – добавляем новый товар
            success = await rq.add_product(product_data)
            if success:
                added_count += 1

    # Получаем все товары из БД и формируем множество их ID
    all_products = await rq.get_all_products()
    db_ids = {product["id"] for product in all_products if product.get("id") is not None}

    # Находим ID товаров, которые есть в БД, но отсутствуют в Excel-файле – их необходимо удалить
    ids_to_delete = db_ids - excel_ids

    for product_id in ids_to_delete:
        if await rq.delete_product(product_id):
            deleted_count += 1

    await state.clear()
    sent_message = await message.answer(
        f"Операция завершена.\nОбновлено товаров: {updated_count}\nДобавлено товаров: {added_count}\nУдалено товаров: {deleted_count}",
        reply_markup=kb.go_to_dashboard
    )
    user_data['bot_messages'].append(sent_message.message_id)


async def get_report_filters(state: FSMContext) -> dict:
    data = await state.get_data()
    if "report_filters" not in data:
        filters = {
            "date": "all",    # варианты: "all", "year", "month", "week" или пользовательский период
            "manager": "all", # "all" или id менеджера
            "status": "all"   # "all", "Ожидание", "Принято", "Отменено"
        }
        await state.update_data(report_filters=filters)
    else:
        filters = data["report_filters"]
    return filters


@router.callback_query(F.data == "view_reports")
async def view_reports(callback_query: CallbackQuery, state: FSMContext):
    tuid = callback_query.message.chat.id
    user_data = sent_message_add_screen_ids[tuid]
    user_data['user_messages'].append(callback_query.message.message_id)
    await delete_previous_messages(callback_query.message, tuid)

    filters = await get_report_filters(state)
    # Получаем текущее значение менеджера
    manager_value = filters.get("manager", "all")
    if manager_value != "all":
        manager = await rq.get_manager_by_id(int(manager_value))
        manager_name = manager["name"] if manager else "все менеджеры"
    else:
        manager_name = "все менеджеры"

    main_keyboard = kb.create_report_main_keyboard(filters, manager_name=manager_name)
    sent_message = await callback_query.message.answer_photo(
        photo=utils.adminka_png,
        caption="Настройте фильтры, нажмите «Создать отчёт», и Excel-документ будет готов.",
        reply_markup=main_keyboard
    )
    user_data['bot_messages'].append(sent_message.message_id)


@router.callback_query(F.data == "filter_date")
async def show_date_filter(callback_query: CallbackQuery, state: FSMContext):
    await callback_query.message.edit_reply_markup(reply_markup=kb.date_filter_keyboard())


@router.callback_query(F.data.startswith("filter_date:"))
async def select_date_filter(callback_query: CallbackQuery, state: FSMContext):

    period = callback_query.data.split(":")[1]
    filters = await get_report_filters(state)
    if period == "period":
        tuid = callback_query.message.chat.id
        user_data = sent_message_add_screen_ids[tuid]
        user_data['user_messages'].append(callback_query.message.message_id)
        await delete_previous_messages(callback_query.message, tuid)

        sent_message = await callback_query.message.answer("Введите период в формате ДД.ММ.ГГГГ - ДД.ММ.ГГГГ")

        user_data['bot_messages'].append(sent_message.message_id)
        await state.set_state(st.ReportFilter.waiting_for_period_input)
        return
    else:
        filters["date"] = period
        await state.update_data(report_filters=filters)

    manager_value = filters.get("manager", "all")
    if manager_value != "all":
        manager = await rq.get_manager_by_id(int(manager_value))
        manager_name = manager["name"] if manager else "все менеджеры"
    else:
        manager_name = "все менеджеры"

    main_keyboard = kb.create_report_main_keyboard(filters, manager_name=manager_name)
    await callback_query.message.edit_reply_markup(reply_markup=main_keyboard)


@router.message(st.ReportFilter.waiting_for_period_input)
async def process_period_input(message: Message, state: FSMContext):
    tuid = message.chat.id
    user_data = sent_message_add_screen_ids[tuid]
    user_data['user_messages'].append(message.message_id)
    await delete_previous_messages(message, tuid)

    period_text = message.text.strip()
    if period_text == "/start":
        await admin_account(message, state)
        return
    else:
        if " - " not in period_text:
            sent_message = await message.answer("Неверный формат. Введите период в формате ДД.ММ.ГГГГ - ДД.ММ.ГГГГ")
            user_data['bot_messages'].append(sent_message.message_id)
            return
    filters = await get_report_filters(state)
    filters["date"] = period_text
    await state.update_data(report_filters=filters)

    manager_value = filters.get("manager", "all")
    if manager_value != "all":
        manager = await rq.get_manager_by_id(int(manager_value))
        manager_name = manager["name"] if manager else "все менеджеры"
    else:
        manager_name = "все менеджеры"

    main_keyboard = kb.create_report_main_keyboard(filters, manager_name=manager_name)
    sent_message = await message.answer_photo(
        photo=utils.adminka_png,
        caption="Настройте фильтры, нажмите «Создать отчёт», и Excel-документ будет готов.",
        reply_markup=main_keyboard
    )
    user_data['bot_messages'].append(sent_message.message_id)


@router.callback_query(F.data == "filter_manager")
async def show_manager_filter(callback_query: CallbackQuery, state: FSMContext):
    managers = await rq.get_all_managers()
    keyboard = kb.manager_filter_keyboard(managers)
    await callback_query.message.edit_reply_markup(reply_markup=keyboard)


@router.callback_query(F.data.startswith("filter_manager:"))
async def select_manager_filter(callback_query: CallbackQuery, state: FSMContext):
    manager_value = callback_query.data.split(":")[1]
    filters = await get_report_filters(state)
    filters["manager"] = manager_value
    await state.update_data(report_filters=filters)

    if manager_value != "all":
        manager = await rq.get_manager_by_id(int(manager_value))
        manager_name = manager["name"] if manager else "все менеджеры"
    else:
        manager_name = "все менеджеры"

    main_keyboard = kb.create_report_main_keyboard(filters, manager_name=manager_name)
    await callback_query.message.edit_reply_markup(reply_markup=main_keyboard)


@router.callback_query(F.data == "filter_status")
async def show_status_filter(callback_query: CallbackQuery, state: FSMContext):
    keyboard = kb.status_filter_keyboard()
    await callback_query.message.edit_reply_markup(reply_markup=keyboard)


@router.callback_query(F.data.startswith("filter_status:"))
async def select_status_filter(callback_query: CallbackQuery, state: FSMContext):
    status_value = callback_query.data.split(":")[1]
    filters = await get_report_filters(state)
    filters["status"] = status_value
    await state.update_data(report_filters=filters)

    manager_value = filters.get("manager", "all")
    if manager_value != "all":
        manager = await rq.get_manager_by_id(int(manager_value))
        manager_name = manager["name"] if manager else "все менеджеры"
    else:
        manager_name = "все менеджеры"

    main_keyboard = kb.create_report_main_keyboard(filters, manager_name=manager_name)
    await callback_query.message.edit_reply_markup(reply_markup=main_keyboard)


@router.callback_query(F.data == "back_to_report_main")
async def back_to_report_main(callback_query: CallbackQuery, state: FSMContext):
    filters = await get_report_filters(state)

    manager_value = filters.get("manager", "all")
    if manager_value != "all":
        manager = await rq.get_manager_by_id(int(manager_value))
        manager_name = manager["name"] if manager else "все менеджеры"
    else:
        manager_name = "все менеджеры"

    main_keyboard = kb.create_report_main_keyboard(filters, manager_name=manager_name)
    await callback_query.message.edit_reply_markup(reply_markup=main_keyboard)


@router.callback_query(F.data == "create_report")
async def create_report(callback_query: CallbackQuery, state: FSMContext):
    tuid = callback_query.message.chat.id
    user_data = sent_message_add_screen_ids[tuid]
    user_data['user_messages'].append(callback_query.message.message_id)
    await delete_previous_messages(callback_query.message, tuid)

    filters = await get_report_filters(state)
    report_file, report_filename = await generate_report_excel(filters)
    from aiogram.types.input_file import BufferedInputFile
    input_file = BufferedInputFile(report_file.getvalue(), filename=report_filename)
    sent_message = await callback_query.message.answer_document(
        document=input_file,
        caption="Отчёт сформирован и готов к скачиванию.",
        reply_markup=kb.go_to_dashboard
    )

    user_data['bot_messages'].append(sent_message.message_id)


@router.callback_query(F.data == "send_notifications")
async def broadcast_start(callback_query: CallbackQuery, state: FSMContext):
    tuid = callback_query.message.chat.id
    user_data = sent_message_add_screen_ids[tuid]
    user_data['user_messages'].append(callback_query.message.message_id)
    await delete_previous_messages(callback_query.message, tuid)

    sent_message = await callback_query.message.answer_photo(
        photo=utils.adminka_png,
        caption="Выберите нужное действие с рассылками:",
        reply_markup=kb.broadcast_menu_start
    )
    user_data['bot_messages'].append(sent_message.message_id)


@router.callback_query(F.data == "broadcast_history")
async def show_broadcast_history(callback_query: CallbackQuery, state: FSMContext):
    tuid = callback_query.message.chat.id
    user_data = sent_message_add_screen_ids[tuid]
    user_data['user_messages'].append(callback_query.message.message_id)
    await delete_previous_messages(callback_query.message, tuid)

    # Теперь функция сама создаёт сессию
    history = await rq.get_broadcast_history()

    group_names = {
        "managers": "менеджеров",
        "leads": "лидов",
        "clients": "клиентов",
        "all": "всех клиентов"
    }

    if history:
        message_text = "ИСТОРИЯ РАССЫЛОК:\n\n"
        for item in history:
            text = item.text[:30] + "..." if len(item.text) > 20 else item.text
            date = item.created_at.strftime("%d.%m.%Y")
            group_name = group_names.get(item.target_group, item.target_group)
            message_text += f"✏️ {text}\n📅 {date} | {item.total_users} • {item.delivered} • {item.failed} | {group_name}\n\n"
    else:
        message_text = "История рассылок пуста."

    sent_message = await callback_query.message.answer(message_text, reply_markup=kb.go_to_notification)
    user_data['bot_messages'].append(sent_message.message_id)


@router.callback_query(F.data == "create_broadcast")
async def show_broadcast_filters(callback_query: CallbackQuery, state: FSMContext):
    tuid = callback_query.message.chat.id
    user_data = sent_message_add_screen_ids[tuid]
    user_data['user_messages'].append(callback_query.message.message_id)
    await delete_previous_messages(callback_query.message, tuid)

    sent_message = await callback_query.message.answer_photo(
        photo=utils.adminka_png,
        caption="Выберите кому нужно отправить рассылки",
        reply_markup=kb.broadcast_filter_menu
    )
    user_data['bot_messages'].append(sent_message.message_id)

    @router.callback_query(F.data.startswith("filter_"))
    async def set_filter_type(callback_query: CallbackQuery, state: FSMContext):
        filter_type = callback_query.data.split("_")[1]
        await state.update_data(filter_type=filter_type)
        await start_broadcast(callback_query, state)


async def start_broadcast(callback_query: CallbackQuery, state: FSMContext):
    tuid = callback_query.message.chat.id
    user_data = sent_message_add_screen_ids[tuid]
    user_data['user_messages'].append(callback_query.message.message_id)
    await delete_previous_messages(callback_query.message, tuid)

    await state.set_state(st.BroadcastStates.waiting_for_media)
    sent_message = await callback_query.message.answer("Добавьте фото/видео для рассылки или отправьте -, если его нет.")
    user_data['bot_messages'].append(sent_message.message_id)


@router.message(st.BroadcastStates.waiting_for_media, F.photo | F.video | F.text)
async def receive_media(message: Message, state: FSMContext):
    tuid = message.chat.id
    user_data = sent_message_add_screen_ids[tuid]
    user_data['user_messages'].append(message.message_id)
    await delete_previous_messages(message, tuid)

    media = None
    if message.photo:
        media = message.photo[-1].file_id
        await state.update_data(media_type="photo")
    elif message.video:
        media = message.video.file_id
        await state.update_data(media_type="video")
    elif message.text and message.text == "-":
        await state.update_data(media=None)
        await state.update_data(media_type=None)
        await state.set_state(st.BroadcastStates.waiting_for_text)
        sent_message = await message.answer("Введите текст для рассылки.")
        user_data['bot_messages'].append(sent_message.message_id)
        return

    await state.update_data(media=media)
    await state.set_state(st.BroadcastStates.waiting_for_text)
    sent_message = await message.answer("Введите текст для рассылки.")
    user_data['bot_messages'].append(sent_message.message_id)


@router.message(st.BroadcastStates.waiting_for_text, F.text | F.caption)
async def receive_text(message: Message, state: FSMContext):
    tuid = message.chat.id
    user_data = sent_message_add_screen_ids[tuid]
    user_data['user_messages'].append(message.message_id)
    await delete_previous_messages(message, tuid)

    if message.text:
        await state.update_data(text=message.text)
    elif message.caption:
        await state.update_data(text=message.caption)

    data = await state.get_data()
    media = data.get("media")
    text = data.get("text")

    async def send_broadcast(media, text, media_type):
        if media:
            if media_type == "photo":
                return InputMediaPhoto(media=media, caption=text)
            elif media_type == "video":
                return InputMediaVideo(media=media, caption=text)
        else:
            return text

    media_object = await send_broadcast(media, text, data.get("media_type"))

    if media_object:
        if isinstance(media_object, str):
            sent_message = await message.answer(f"Вот так будет выглядеть рассылка:\n\n{media_object}",
                                 reply_markup=kb.broadcast_menu)
            user_data['bot_messages'].append(sent_message.message_id)
        else:
            if isinstance(media_object, InputMediaPhoto):
                sent_message = await message.answer_photo(media_object.media, caption=media_object.caption,
                                           reply_markup=kb.broadcast_menu)
                user_data['bot_messages'].append(sent_message.message_id)
            elif isinstance(media_object, InputMediaVideo):
                sent_message = await message.answer_video(media_object.media, caption=media_object.caption,
                                           reply_markup=kb.broadcast_menu)
                user_data['bot_messages'].append(sent_message.message_id)
    else:
        sent_message = await message.answer("Вот так будет выглядеть рассылка:\n\n" + text, reply_markup=kb.broadcast_menu)
        user_data['bot_messages'].append(sent_message.message_id)

    await state.set_state(st.BroadcastStates.confirmation)


@router.callback_query(F.data == "broadcast_confirm")
async def confirm_broadcast(callback_query: CallbackQuery, state: FSMContext, data=None):
    tuid = callback_query.message.chat.id
    user_data = sent_message_add_screen_ids[tuid]
    user_data['user_messages'].append(callback_query.message.message_id)
    await delete_previous_messages(callback_query.message, tuid)

    data = await state.get_data()
    users = await rq.get_all_users()
    media = data.get("media")
    text = data.get("text")
    media_type = data.get("media_type")
    filter_type = data.get("filter_type")

    if filter_type:
        users = await rq.get_filtered_users(filter_type)
    else:
        users = await rq.get_filtered_users("all_clients")
    total_users = len(users)
    delivered = 0
    failed = 0

    for user_id in users:
        try:
            if media:
                if media_type == "photo":
                    await callback_query.bot.send_photo(user_id, media, caption=text, reply_markup=kb.close_broadcast_message)
                elif media_type == "video":
                    await callback_query.bot.send_video(user_id, media, caption=text, reply_markup=kb.close_broadcast_message)
            else:
                await callback_query.bot.send_message(user_id, text, reply_markup=kb.close_broadcast_message)
            delivered += 1
        except Exception as e:
            print(f"Ошибка отправки пользователю {user_id}: {e}")
            failed += 1

    # Сохраняем историю рассылки через функцию, внутри которой создаётся сессия
    await rq.save_broadcast_history(
         text=text,
         media_type=media_type,
         media_file_id=media,
         total_users=total_users,
         delivered=delivered,
         failed=failed,
         target_group=filter_type if filter_type else "all_clients"
    )

    sent_message = await callback_query.message.answer_photo(
        photo=utils.adminka_png,
        caption="Рассылка успешно отправлена! Возвращаемся в главное меню.",
        reply_markup=kb.go_to_dashboard
    )
    user_data['bot_messages'].append(sent_message.message_id)
    await state.clear()


@router.callback_query(F.data == "broadcast_edit")
async def edit_broadcast(callback_query: CallbackQuery, state: FSMContext):
    tuid = callback_query.message.chat.id
    user_data = sent_message_add_screen_ids[tuid]
    user_data['user_messages'].append(callback_query.message.message_id)
    await delete_previous_messages(callback_query.message, tuid)

    await state.set_state(st.BroadcastStates.waiting_for_media)
    sent_message = await callback_query.message.answer("Добавьте фото/видео для рассылки или отправьте -, если его нет.")
    user_data['bot_messages'].append(sent_message.message_id)


@router.callback_query(F.data == "broadcast_cancel")
async def cancel_broadcast(callback_query: CallbackQuery, state: FSMContext):
    tuid = callback_query.message.chat.id
    user_data = sent_message_add_screen_ids[tuid]
    user_data['user_messages'].append(callback_query.message.message_id)
    await delete_previous_messages(callback_query.message, tuid)

    sent_message = await callback_query.message.answer_photo(
        photo=utils.adminka_png,
        caption="Рассылка отменена. Возвращаемся в главное меню.",
        reply_markup=kb.go_to_dashboard
    )
    user_data['bot_messages'].append(sent_message.message_id)
    await state.clear()


@router.callback_query(F.data == "close_broadcast_message")
async def close_broadcast_message_handler(callback_query: CallbackQuery, state: FSMContext):
    await callback_query.message.delete()


# Обработчик нажатия кнопки "Сгенерировать id фото"
@router.callback_query(F.data == 'generate_photo_id')
async def cmd_generate_photo_id(callback_query: CallbackQuery, state: FSMContext):
    tuid = callback_query.message.chat.id
    user_data = sent_message_add_screen_ids[tuid]
    user_data['user_messages'].append(callback_query.message.message_id)
    await delete_previous_messages(callback_query.message, tuid)

    sent_message = await callback_query.message.answer_photo(
        photo=utils.user_second_png,
        caption="Пожалуйста, отправьте ZIP архив с фотографиями товаров.",
        reply_markup=kb.go_to_dashboard
    )
    user_data['bot_messages'].append(sent_message.message_id)
    await state.set_state(st.PhotoIdGen.waiting_for_zip)

# Обработчик получения ZIP файла (с фильтром по состоянию)
@router.message(
    st.PhotoIdGen.waiting_for_zip,
    lambda message: message.document is not None and message.document.file_name.lower().endswith('.zip')
)
async def handle_zip_file(message: Message, state: FSMContext):
    tuid = message.chat.id
    user_data = sent_message_add_screen_ids[tuid]
    user_data['user_messages'].append(message.message_id)
    await delete_previous_messages(message, tuid)

    sent_message = await message.answer("Обработка архива, пожалуйста, подождите...")

    user_data['bot_messages'].append(sent_message.message_id)
    try:
        # Вызов логики из scripts.py: распаковка архива, отправка фото и генерация PDF
        pdf_file_path = await generate_photo_id_pdf(message)
    except Exception as e:
        await message.answer("Ошибка при обработке архива: " + str(e))
        await state.clear()
        return

    await delete_previous_messages(message, tuid)

    # Отправляем PDF пользователю
    await message.answer_document(
        document=FSInputFile(pdf_file_path),
        caption="Вот ваш PDF файл с file_id фотографий.",
        reply_markup=kb.go_to_dashboard
    )
    user_data['bot_messages'].append(sent_message.message_id)

    await state.clear()

# Обработчик, если прислан не ZIP архив
@router.message(st.PhotoIdGen.waiting_for_zip)
async def invalid_file(message: Message, state: FSMContext):
    tuid = message.chat.id
    user_data = sent_message_add_screen_ids[tuid]
    user_data['user_messages'].append(message.message_id)
    await delete_previous_messages(message, tuid)

    sent_message = await message.answer("Пожалуйста, отправьте архив в формате .zip")

    user_data['bot_messages'].append(sent_message.message_id)



